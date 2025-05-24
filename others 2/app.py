import streamlit as st
import cv2
import os
import pickle
import logging
from deepface import DeepFace
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("system_activity.log")  # Save logs to a file for activity tracking
    ]
)

# File configurations
FACE_DB_FILE = "faces.pkl"
FACE_DIR = "ImageAttendance"
ATTENDANCE_EXCEL_FILE = "attendance.xlsx"

# Create necessary directories and files if they don't exist
os.makedirs(FACE_DIR, exist_ok=True)

# Initialize the Excel file if it doesn't exist
if not os.path.exists(ATTENDANCE_EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time"])  # Add headers
    wb.save(ATTENDANCE_EXCEL_FILE)

# Functions for saving and loading face data
def save_faces(faces):
    """Save faces database."""
    try:
        with open(FACE_DB_FILE, 'wb') as f:
            pickle.dump(faces, f)
        logging.info(f"Database saved with {len(faces)} faces.")
    except Exception as e:
        logging.error(f"Error saving faces database: {e}")

def load_faces():
    """Load faces database or create a new one."""
    if os.path.exists(FACE_DB_FILE):
        try:
            with open(FACE_DB_FILE, 'rb') as f:
                faces = pickle.load(f)
            logging.info(f"Loaded {len(faces)} faces from database.")
            return faces
        except (FileNotFoundError, pickle.PickleError) as e:
            logging.error(f"Error loading faces database: {e}")
    return {}

def detect_face(frame):
    """Detect faces in a frame."""
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    return face_cascade.detectMultiScale(gray, 1.1, 4)

def mark_attendance(name):
    """Mark attendance in the Excel file and ensure it is linked to the Streamlit portal."""
    now = datetime.now()
    date_string = now.strftime('%Y-%m-%d')
    time_string = now.strftime('%H:%M:%S')

    try:
        # Load the Excel file
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active
    except Exception as e:
        # If the file cannot be loaded, reinitialize it
        logging.error(f"Error loading Excel file: {e}. Reinitializing file.")
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time"])  # Add headers

    try:
        # Append the attendance record
        ws.append([name, date_string, time_string])
        wb.save(ATTENDANCE_EXCEL_FILE)
        logging.info(f"Attendance marked for {name} at {date_string} {time_string}.")
        return True
    except Exception as e:
        logging.error(f"Error marking attendance: {e}")
        return False

def register_face():
    """Register a new face using the webcam."""
    name = st.text_input("Enter person's name:")
    if st.button("Register"):
        if not name:
            st.error("Name cannot be empty.")
            return

        faces_db = load_faces()
        person_id = name.lower().replace(" ", "_")
        if person_id in faces_db:
            st.warning(f"Face for {name} is already registered.")
            return

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            st.error("Unable to access the webcam. Please check your camera and try again.")
            return

        st.info("Press 'c' to capture or 'q' to quit.")
        while True:
            ret, frame = cap.read()
            if not ret:
                st.error("Failed to capture frame.")
                break

            faces = detect_face(frame)
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.imshow("Register Face", frame)

            key = cv2.waitKey(1) & 0xFF
            if key == ord('c'):
                if len(faces) > 0:
                    x, y, w, h = faces[0]
                    face_crop = frame[y:y+h, x:x+w]
                    face_path = os.path.join(FACE_DIR, f"{person_id}.jpg")
                    cv2.imwrite(face_path, face_crop)

                    faces_db[person_id] = {"name": name, "path": face_path}
                    save_faces(faces_db)
                    st.success(f"{name} registered successfully.")
                else:
                    st.warning("No face detected. Please try again.")
                break
            elif key == ord('q'):
                st.info("Registration canceled.")
                break

        cap.release()
        cv2.destroyAllWindows()

def verify_face():
    """Verify a face and mark attendance."""
    faces = load_faces()
    if not faces:
        st.info("No faces registered yet. Please register a face first.")
        return

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        st.error("Unable to access the webcam. Please check your camera and try again.")
        return

    st.info("Press 'v' to verify or 'q' to quit.")
    while True:
        ret, frame = cap.read()
        if not ret:
            st.error("Failed to capture frame from webcam.")
            break

        face_rects = detect_face(frame)
        for (x, y, w, h) in face_rects:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

        cv2.imshow("Verify Face", frame)

        key = cv2.waitKey(1) & 0xFF
        if key == ord('v'):
            if len(face_rects) == 0:
                st.info("No face detected. Please try again.")
                continue

            temp_path = "temp.jpg"
            try:
                cv2.imwrite(temp_path, frame)

                match_found = False
                best_match = None
                best_confidence = float('inf')
                CONFIDENCE_THRESHOLD = 0.55

                for person_id, data in faces.items():
                    if data["path"] == "temp.jpg":
                        continue
                    try:
                        result = DeepFace.verify(
                            img1_path=temp_path,
                            img2_path=data["path"],
                            detector_backend='opencv'
                        )
                        if result.get("verified", False):
                            confidence = result.get("distance", float('inf'))
                            if confidence < CONFIDENCE_THRESHOLD and confidence < best_confidence:
                                best_confidence = confidence
                                best_match = data["name"]
                                match_found = True
                    except Exception as e:
                        logging.error(f"Error verifying face with {data['name']}: {e}")

                if match_found and best_match:
                    success = mark_attendance(best_match)
                    if success:
                        st.success(f"Attendance successfully recorded for {best_match}.")
                    else:
                        st.error(f"Failed to record attendance for {best_match}.")
                else:
                    st.info("No match found. Attendance not recorded.")
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            break
        elif key == ord('q'):
            st.info("Verification canceled.")
            break

    cap.release()
    cv2.destroyAllWindows()

# Streamlit UI
st.title("Face Recognition Attendance System")

menu = ["Register Face", "Verify Face", "View Attendance", "List Faces", "Admin Panel"]
choice = st.sidebar.selectbox("Menu", menu)

if choice == "Register Face":
    register_face()
elif choice == "Verify Face":
    verify_face()
elif choice == "View Attendance":
    st.subheader("Attendance Records")
    try:
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        st.table(data)
    except FileNotFoundError:
        st.error("Attendance file not found.")
elif choice == "List Faces":
    st.subheader("Registered Faces")
    faces = load_faces()
    if faces:
        for person_id, details in faces.items():
            st.write(f"{details['name']} (Path: {details['path']})")
    else:
        st.info("No faces registered yet.")
elif choice == "Admin Panel":
    st.subheader("Admin Panel")

    # Admin Authentication
    admin_password = "admin123"  # Set a secure password here
    password = st.text_input("Enter Admin Password", type="password")
    if st.button("Login"):
        if password == admin_password:
            st.success("Access granted to Admin Panel.")
            
            admin_action = st.selectbox("Choose an action", ["Add User", "Remove User", "Update User Profile"])

            if admin_action == "Add User":
                st.subheader("Add a New User")
                register_face()

            elif admin_action == "Remove User":
                st.subheader("Remove a User")
                faces = load_faces()
                if not faces:
                    st.info("No users registered yet.")
                else:
                    user_to_remove = st.selectbox("Select a user to remove", [details["name"] for details in faces.values()])
                    if st.button("Remove"):
                        for person_id, details in list(faces.items()):
                            if details["name"] == user_to_remove:
                                del faces[person_id]
                                save_faces(faces)
                                st.success(f"User '{user_to_remove}' removed successfully.")
                                break

            elif admin_action == "Update User Profile":
                st.subheader("Update User Profile")
                faces = load_faces()
                if not faces:
                    st.info("No users registered yet.")
                else:
                    user_to_update = st.selectbox("Select a user to update", [details["name"] for details in faces.values()])
                    new_name = st.text_input("Enter the new name")
                    if st.button("Update"):
                        for person_id, details in faces.items():
                            if details["name"] == user_to_update:
                                details["name"] = new_name
                                save_faces(faces)
                                st.success(f"User '{user_to_update}' updated to '{new_name}' successfully.")
                                break
        else:
            st.error("Access denied. Incorrect password.")