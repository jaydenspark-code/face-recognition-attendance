import cv2
import face_recognition
import pickle
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

DB_FILE = "face_db.pkl"
ATTENDANCE_FILE = "attendance.xlsx"

def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "rb") as f:
            return pickle.load(f)
    return {}

def save_db(db):
    with open(DB_FILE, "wb") as f:
        pickle.dump(db, f)

def init_attendance_file():
    if not os.path.exists(ATTENDANCE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time"])
        wb.save(ATTENDANCE_FILE)

def register_user():
    name = input("Enter your name: ").strip()
    db = load_db()
    if name in db:
        print("User already registered.")
        return

    cap = cv2.VideoCapture(0)
    print("Press SPACE to capture your face.")
    while True:
        ret, frame = cap.read()
        cv2.imshow("Register Face", frame)
        if cv2.waitKey(1) & 0xFF == ord(' '):
            break
    cap.release()
    cv2.destroyAllWindows()

    face_locations = face_recognition.face_locations(frame)
    if not face_locations:
        print("No face detected. Try again.")
        return
    face_encoding = face_recognition.face_encodings(frame, face_locations)[0]
    db[name] = face_encoding
    save_db(db)
    print(f"User '{name}' registered successfully.")

def mark_attendance():
    db = load_db()
    if not db:
        print("No users registered yet.")
        return

    cap = cv2.VideoCapture(0)
    print("Press SPACE to capture your face for attendance.")
    while True:
        ret, frame = cap.read()
        cv2.imshow("Attendance", frame)
        if cv2.waitKey(1) & 0xFF == ord(' '):
            break
    cap.release()
    cv2.destroyAllWindows()

    face_locations = face_recognition.face_locations(frame)
    if not face_locations:
        print("No face detected. Try again.")
        return
    face_encoding = face_recognition.face_encodings(frame, face_locations)[0]

    for name, db_encoding in db.items():
        match = face_recognition.compare_faces([db_encoding], face_encoding)[0]
        if match:
            print(f"Attendance marked for {name}!")
            log_attendance(name)
            return
    print("Face not recognized. Attendance not marked.")

def log_attendance(name):
    init_attendance_file()
    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active
    now = datetime.now()
    ws.append([name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
    wb.save(ATTENDANCE_FILE)
    print("Attendance logged in Excel.")

if __name__ == "__main__":
    print("1. Register New User")
    print("2. Mark Attendance")
    choice = input("Select option: ")
    if choice == "1":
        register_user()
    elif choice == "2":
        mark_attendance()
    else:
        print("Invalid option.")