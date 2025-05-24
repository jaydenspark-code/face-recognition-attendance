import cv2
import os
import pickle
import logging
import streamlit as st
import numpy as np
from deepface import DeepFace
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import time
import os
import logging
import dlib
from scipy.spatial import distance as dist
from deepface import DeepFace
from numpy.linalg import norm

def cosine_similarity(a, b):
    return np.dot(a, b) / (norm(a) * norm(b))


# --- Configuration ---
LOG_FILE = "system_activity.log"
FACE_DB_FILE = "faces.pkl"
FACE_DIR = "ImageAttendance"
ATTENDANCE_EXCEL_FILE = "attendance.xlsx"
USER_DB_FILE = "users.pkl"
ADMIN_PASSWORD = "admin123"  # IMPORTANT: Use a more secure password in production

LECTURER_COURSE_MAPPING = {
    "Office Productivity Software Tools": {
        "lecturer": "Engr. Victoria Dansowaa",
         "email": None  # Email not available yet
    },
    "Foreign Language (French) I": {
        "lecturer": "Mr. Isaac Kofi Evan Danjo",
         "email": None  # Email not available yet
    },
    "Programming for Problem Solving": {
        "lecturer": "Mr. Nana Owusu Marfo",
        "email": None  # Email not available yet
    },
    "Intro. to Cyber Security and Digital Forensics": {
        "lecturer": "Dr. Vivek",
         "email": None  # Email not available yet
    },
    "Communication Skills in English I": {
        "lecturer": "Ms. Jane Smith",
         "email": None  # Email not available yet
    },
    "Linear Algebra": {
        "lecturer": "Mr. Enoch Grant",
         "email": None  # Email not available yet
    },
    "African Studies": {
        "lecturer": "Mr. Samuel Welbeck",
         "email": None  # Email not available yet
    }
}


LECTURER_DB_FILE = "lecturers.pkl"  # File to store lecturer data

def save_lecturers(data):
    """Save LECTURER_COURSE_MAPPING to a file."""
    try:
        with open(LECTURER_DB_FILE, 'wb') as f:
            pickle.dump(data, f)
        logging.info(f"Lecturer data saved to {LECTURER_DB_FILE}")
    except Exception as e:
        logging.error(f"Error saving lecturer data: {e}")

def load_lecturers():
    """Load LECTURER_COURSE_MAPPING from a file."""
    if os.path.exists(LECTURER_DB_FILE):
        try:
            with open(LECTURER_DB_FILE, 'rb') as f:
                data = pickle.load(f)
            logging.info(f"Lecturer data loaded from {LECTURER_DB_FILE}")
            return data
        except Exception as e:
            logging.error(f"Error loading lecturer data: {e}")
    return {}  # Return an empty dictionary if the file does not exist

COURSE_TIMETABLE = {
    "Monday": {"13:00-15:30": {"course": "Office Productivity Software Tools", "code": "BNCS111"}},
    "Tuesday": {
        "13:30-15:30": {"course": "Foreign Language (French) I", "code": "GDP005"},
        "16:00-19:00": {"course": "Programming for Problem Solving", "code": "BNCS113"},
    },
    "Wednesday": {"10:00-13:00": {"course": "Intro. to Cyber Security and Digital Forensics", "code": "BNCS109"}},
    "Thursday": {
        "10:00-12:00": {"course": "Communication Skills in English I", "code": "GDP001"},
        "13:30-16:30": {"course": "Linear Algebra", "code": "MAT001"},
    },
    "Friday": {"10:00-12:00": {"course": "African Studies", "code": "GPD003"}},
}

CONFIRMATION_LOG_FILE = "confirmation_logs.xlsx"

# Ensure the confirmation log file exists
if not os.path.exists(CONFIRMATION_LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Course", "Date", "Time", "Course Representative", "Role", "Lecturer", "Timestamp"])
    wb.save(CONFIRMATION_LOG_FILE)

# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler(LOG_FILE)],
)

# --- File Management ---
os.makedirs(FACE_DIR, exist_ok=True)
if not os.path.exists(ATTENDANCE_EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time", "Course", "Verified By"])
    wb.save(ATTENDANCE_EXCEL_FILE)

# --- Data Persistence ---
def save_data(data, filename):
    """Save data to a pickle file."""
    try:
        with open(filename, 'wb') as f:
            pickle.dump(data, f)
        logging.info(f"Data saved to {filename}")
    except Exception as e:
        logging.error(f"Error saving data to {filename}: {e}")

def load_data(filename, default=None):
    """Load data from a pickle file or return default if not found."""
    if os.path.exists(filename):
        try:
            with open(filename, 'rb') as f:
                data = pickle.load(f)
            logging.info(f"Data loaded from {filename}")
            return data
        except Exception as e:
            logging.error(f"Error loading data from {filename}: {e}")
    return default if default is not None else {}

# --- Face Detection ---
def detect_face(frame):
    """Detect faces in a frame."""
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    return face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

def eye_aspect_ratio(eye):
    A = dist.euclidean(eye[1], eye[5])
    B = dist.euclidean(eye[2], eye[4])
    C = dist.euclidean(eye[0], eye[3])
    ear = (A + B) / (2.0 * C)
    return ear

def is_blurry(image, threshold=100):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    fm = cv2.Laplacian(gray, cv2.CV_64F).var()
    return fm < threshold

# --- Email Notification ---
def send_email_notification(student_name, course_title, student_email):
    """Send an email notification to a student when their attendance is marked."""
    sender_email = "thearnest7@gmail.com"
    sender_password = "eutzqmfzvrqqdxbi"  # Use your Gmail app password here

    if not sender_email or not sender_password:
        logging.error("Email credentials are not fully set (sender email or password).")
        # If using Streamlit, uncomment the next line:
        # st.error("Email credentials are not configured. Please contact the administrator.")
        return False

    subject = f"Attendance Marked for {course_title}"
    body = f"Dear {student_name},\n\nYour attendance has been successfully marked for the course: {course_title}.\n\nBest regards,\nAttendance System"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = student_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))


    for attempt in range(3):  # Retry up to 3 times
        try:
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            server.login(sender_email, sender_password)
            server.send_message(msg)
            server.quit()
            logging.info(f"Attendance notification sent to {student_email} for {student_name} in {course_title}.")
            return True
        except smtplib.SMTPException as e:
            logging.error(f"Attempt {attempt + 1}: Failed to send email: {e}")
            if attempt < 2:  # Wait before retrying
                time.sleep(5)  # Wait 5 seconds before retrying
            else:
                # If using Streamlit, uncomment the next line:
                # st.error(f"An error occurred while sending the email notification: {e}")
                print(f"An error occurred while sending the email notification: {e}")
                return False
    
def generate_and_send_summary(course_title):
    """Generate attendance summary and send it to admin and lecturer."""
    try:
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active
        attendance_list = [
            f"{row[0]} - {row[1]} {row[2]}"  # Name, Date, Time
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[3] == course_title  # Filter by course
        ]

        # Send to admin
        admin_email = "thearnest7@gmail.com"
        send_summary_email(course_title, attendance_list, admin_email)

        # Send to lecturer
        lecturer_info = LECTURER_COURSE_MAPPING.get(course_title, {})
        lecturer_email = lecturer_info.get("email")
        if lecturer_email:
            send_summary_email(course_title, attendance_list, lecturer_email)

    except Exception as e:
        logging.error(f"Error generating or sending summary: {e}")

# --- Attendance Marking ---
def mark_attendance(name, course_title):
    """Mark attendance and send email notification."""
    now = datetime.now()
    date_string = now.strftime('%Y-%m-%d')
    time_string = now.strftime('%H:%M:%S')

    try:
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active
    except Exception as e:
        logging.error(f"Error loading Excel file: {e}. Reinitializing file.")
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time", "Course", "Verified By"])

    try:
        ws.append([name, date_string, time_string, course_title, ""])  # Initially not verified
        wb.save(ATTENDANCE_EXCEL_FILE)
        logging.info(f"Attendance marked for {name} in {course_title}.")

        # Retrieve the student's email address from the database
        faces_db = load_data(FACE_DB_FILE, default={})
        student_email = None
        for person_id, details in faces_db.items():
            if details["name"] == name:
                student_email = details.get("email")
                break

        # Attempt to send email notification
        if student_email:
            email_success = send_email_notification(name, course_title, student_email)
            if email_success:
                logging.info(f"Email notification successfully sent for {name}.")
            else:
                logging.warning(f"Attendance marked, but email notification failed for {name}.")
                st.warning(f"Attendance marked, but email notification failed for {name}.")
        else:
            logging.warning(f"Email address not found for {name}. Attendance marked without email notification.")
            st.warning(f"Attendance marked, but no email address found for {name}.")

        return True
    except Exception as e:
        logging.error(f"Error marking attendance: {e}")
        return False
    
def view_attendance():
    """View attendance records from the Excel file."""
    st.subheader("View Attendance Records")
    try:
        # Load the attendance Excel file
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active

        # Extract attendance data
        attendance_data = [
            dict(zip([cell.value for cell in ws[1]], row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

        if not attendance_data:
            st.info("No attendance records found.")
            return

        # Display attendance records in a table
        st.write("### Attendance Records")
        st.dataframe(attendance_data)

    except FileNotFoundError:
        st.error("Attendance file not found.")
    except Exception as e:
        st.error(f"Error loading attendance records: {e}")       
def auto_adjust_brightness(image):
    """Automatically adjust brightness and contrast using CLAHE."""
    lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    cl = clahe.apply(l)
    limg = cv2.merge((cl, a, b))
    adjusted = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)
    return adjusted

def verify_face(course_title):
    """Verify a face and mark attendance using Streamlit's camera input."""
    faces_db = load_data(FACE_DB_FILE, default={})
    if not faces_db:
        st.info("No faces registered yet. Please register a face first.")
        return

    st.info("Use the camera below to capture your face for verification.")
    captured_image = st.camera_input("Capture your face")
    verify_button = st.button("Verify Face")

    face_crop = None
    if captured_image:
        bytes_data = captured_image.getvalue()
        frame = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
        frame = auto_adjust_brightness(frame)

        faces = detect_face(frame)
        if len(faces) != 1:
            st.warning("Please ensure only one face is visible in the frame.")
            return
        x, y, w, h = faces[0]
        face_crop = frame[y:y + h, x:x + w]
        if is_blurry(face_crop):
            st.warning("Image is too blurry. Please retake the photo.")
            return

        # Draw rectangle and show image
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
        st.image(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB), caption="Detected Face")

    if verify_button:
        if not captured_image or face_crop is None:
            st.warning("No image captured. Please try again.")
            return

        try:
            embedding = DeepFace.represent(
                face_crop,
                model_name="ArcFace",
                detector_backend="retinaface",
                enforce_detection=True,
                align=True,
                anti_spoofing=False
            )[0]["embedding"]
        except ValueError as ve:
            st.warning(f"Verification failed: {ve}")
            return
        except Exception as e:
            st.error(f"An unexpected error occurred during verification: {e}")
            return

        # --- Matching logic ---
        best_match = None
        best_score = -1  # Cosine similarity: higher is better
        for person_id, data in faces_db.items():
            for emb in data.get("embeddings", []):
                score = cosine_similarity(embedding, emb)
                if score > best_score:
                    best_score = score
                    best_match = data["name"]
        # --- Set a threshold for match ---
        if best_score > 0.5:  # You may need to tune this threshold
            success = mark_attendance(best_match, course_title)
            if success:
                st.success(f"Attendance successfully marked for {best_match} in {course_title}.")
            else:
                st.error(f"Failed to mark attendance for {best_match}. Please try again.")
        else:
            st.warning("No matching face found. Attendance not recorded.")
def register_face():
    st.subheader("Register New Face (Multi-Frame)")
    name = st.text_input("Enter person's name:").strip()
    email = st.text_input("Enter person's email address:").strip()

    if not name or not email:
        st.warning("Please provide both the name and email address.")
        return

    person_id = name.lower().replace(" ", "_")
    faces_db = load_data(FACE_DB_FILE, default={})
    if person_id in faces_db:
        st.warning(f"Face for '{name}' is already registered.")
        return

    st.info("Click 'Take Photo' each time you change your head position. Capture at least 5–6 photos from different angles before clicking 'Finish Registration'. You can also register with different looks (with/without glasses, different hairstyles, etc.)")

    if "multi_face_images" not in st.session_state:
        st.session_state.multi_face_images = []

    captured_image = st.camera_input("Capture multiple frames (move your head slowly)", key="multi_frame", help="Move your head and click 'Capture' several times.")

    if captured_image:
        bytes_data = captured_image.getvalue()
        frame = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
        faces = detect_face(frame)
        if len(faces) == 1:
            x, y, w, h = faces[0]
            face_crop = frame[y:y + h, x:x + w]
            st.session_state.multi_face_images.append(face_crop)
            st.image(face_crop, channels="BGR", caption=f"Captured Face #{len(st.session_state.multi_face_images)}")
        else:
            st.warning("Please ensure only one face is visible in the frame.")

    if st.button("Finish Registration"):
        if len(st.session_state.multi_face_images) < 5:
            st.warning("Please capture at least 5 different face images.")
            return
        face_paths = []
        face_embeddings = []
        os.makedirs(FACE_DIR, exist_ok=True)
        for idx, face_img in enumerate(st.session_state.multi_face_images):
            face_path = os.path.join(FACE_DIR, f"{person_id}_{idx}.jpg")
            cv2.imwrite(face_path, face_img)
            face_paths.append(face_path)
            try:
                embedding = DeepFace.represent(
                    face_img,
                    model_name="ArcFace",
                    detector_backend="retinaface",
                    enforce_detection=True,
                    align=True,
                    anti_spoofing=False
                )[0]["embedding"]
                face_embeddings.append(embedding)
            except ValueError as ve:
                st.warning(f"Face image #{idx+1} was rejected: {ve}")
            except Exception as e:
                st.error(f"An unexpected error occurred for image #{idx+1}: {e}")
        faces_db[person_id] = {
            "name": name,
            "email": email,
            "paths": face_paths,
            "embeddings": face_embeddings
        }
        save_data(faces_db, FACE_DB_FILE)
        st.success(f"Multi-frame registration for '{name}' completed!")
        st.session_state.multi_face_images = []
def list_faces():
    """List all registered faces."""
    st.subheader("Registered Faces")
    faces_db = load_data(FACE_DB_FILE, default={})
    if not faces_db:
        st.info("No faces registered yet.")
        return

    # Display the list of registered faces
    for person_id, details in faces_db.items():
        st.write(f"- **Name:** {details['name']} | **ID:** {person_id}")
        # Show all images for this user
        if "paths" in details:
            for idx, img_path in enumerate(details["paths"]):
                if os.path.exists(img_path):
                    st.image(img_path, caption=f"{details['name']}'s Face #{idx+1}", width=150)
                else:
                    st.warning(f"Image {img_path} for {details['name']} not found.")
        elif "path" in details:
            # For backward compatibility with old single-image entries
            if os.path.exists(details["path"]):
                st.image(details["path"], caption=f"{details['name']}'s Face", width=150)
            else:
                st.warning(f"Image for {details['name']} not found.")
# --- Daily Attendance Summary ---
def generate_daily_summary():
    """Generate and display a daily attendance summary with absences based on the timetable."""
    st.subheader("Daily Attendance Summary")
    selected_date = st.date_input("Select Date for Summary", datetime.now().date())
    summary_data = {}
    all_registered_faces = load_data(FACE_DB_FILE)
    all_students = {details['name'] for details in all_registered_faces.values()}

    try:
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active

        # Extract attendance records
        attendance_records = [
            dict(zip([cell.value for cell in ws[1]], row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

        # Filter records for the selected date
        daily_records = [
            record for record in attendance_records
            if record.get("Date") == selected_date.strftime('%Y-%m-%d')
        ] 

        if not daily_records:
            st.info(f"No attendance records found for {selected_date.strftime('%Y-%m-%d')}.")
            return


        day_of_week = selected_date.strftime('%A')
        scheduled_classes = COURSE_TIMETABLE.get(day_of_week, {})

        for time_slot, class_info in scheduled_classes.items():
            course = class_info['course']
            present_students = {
                record['Name'] for record in daily_records if record.get('Course Title') == course
            }
            absent_students = all_students - present_students

            if course not in summary_data:
                summary_data[course] = {"present": set(), "absent": set()}
            summary_data[course]["present"].update(present_students)
            summary_data[course]["absent"].update(absent_students)

        if summary_data:
            st.markdown(f"### Attendance Summary for {selected_date.strftime('%Y-%m-%d')} ({day_of_week})")
            for course, attendance in summary_data.items():
                st.markdown(f"#### **Course:** {course}")
                if attendance["present"]:
                    st.markdown(f"**Present:**")
                    st.markdown(f"- {', '.join(sorted(list(attendance['present'])))}")
                else:
                    st.markdown("**Present:** No students marked present.")
                if attendance["absent"]:
                    st.markdown(f"**Absent:**")
                    st.markdown(f"- {', '.join(sorted(list(attendance['absent'])))}")
                else:
                    st.markdown("**Absent:** All registered students marked present (or no registered students).")
                st.markdown("---")  # Add a horizontal line for separation
        else:
            st.info(f"No scheduled classes or attendance records found for {selected_date.strftime('%Y-%m-%d')}.")

    except FileNotFoundError:
        st.error("Attendance file not found.")
    except Exception as e:
        st.error(f"Error generating daily summary: {e}")

def log_confirmation(course, date, time, rep_name, role, lecturer_name):
    """Log the confirmation details to the confirmation log file."""
    try:
        # Load the confirmation log file
        wb = load_workbook(CONFIRMATION_LOG_FILE)
        ws = wb.active

        # Append the confirmation details
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([course, date, time, rep_name, role, lecturer_name, timestamp])
        wb.save(CONFIRMATION_LOG_FILE)

        logging.info(f"Confirmation logged for {course} on {date} by {lecturer_name} and {rep_name}.")
    except Exception as e:
        logging.error(f"Error logging confirmation: {e}")        

# --- Attendance Verification ---
def verify_attendance():
    """Verify attendance records for a specific course and date."""
    st.subheader("Verify Attendance Records")

    # Initialize session state variables
    if "step" not in st.session_state:
        st.session_state.step = 1  # Start at Step 1
    if "rep_name" not in st.session_state:
        st.session_state.rep_name = None
    if "selected_role" not in st.session_state:
        st.session_state.selected_role = None
    if "verification_success" not in st.session_state:
        st.session_state.verification_success = False

    try:
        # Load the users database
        users = load_data(USER_DB_FILE, {"Lecturers": {}, "CourseReps": {}})

        # Load the attendance Excel file
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active

        # Select a course and date for verification
        courses = list(LECTURER_COURSE_MAPPING.keys())
        selected_course = st.selectbox("Select Course", courses)
        selected_date = st.date_input("Select Date", datetime.now().date())

        # Filter attendance records for the selected course and date
        attendance_records = [
            dict(zip([cell.value for cell in ws[1]], row))
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[3] == selected_course and row[1] == selected_date.strftime('%Y-%m-%d')
        ]

        if not attendance_records:
            st.info(f"No attendance records found for {selected_course} on {selected_date.strftime('%Y-%m-%d')}.")
            return

        # Display attendance records for verification
        st.write(f"### Attendance Records for {selected_course} on {selected_date.strftime('%Y-%m-%d')}")
        st.dataframe(attendance_records)

        # Step 1: Course Representative Confirmation
        if st.session_state.step == 1:
            st.markdown("#### Step 1: Course Representative Confirmation")
            course_reps = users["CourseReps"].get(selected_course, [])
            if not course_reps:
                st.warning("No course representatives found for this course.")
                return

            # Select role (Head or Assistant Course Rep)
            roles = ["Head Course Rep", "Assistant Course Rep"]
            selected_role = st.selectbox("Select Role", roles)
            rep_name = st.text_input(f"Enter {selected_role}'s Name").strip()

            # Validate the entered name against the selected role
            valid_rep = any(rep["name"] == rep_name and rep["role"] == selected_role for rep in course_reps)
            if st.button("Confirm as Course Representative"):
                if not valid_rep:
                    st.warning(f"The name does not match any {selected_role} for this course. Please try again.")
                    return
                st.success(f"Attendance confirmed by {selected_role}: {rep_name}")
                st.session_state.step = 2  # Move to Step 2
                st.session_state.rep_name = rep_name
                st.session_state.selected_role = selected_role

        # Step 2: Lecturer Verification
        if st.session_state.step == 2:
            st.markdown("#### Step 2: Lecturer Verification")
            lecturer_info = LECTURER_COURSE_MAPPING.get(selected_course, {})
            lecturer_name = st.text_input("Enter Lecturer's Name").strip()

            if st.button("Verify as Lecturer"):
                # Validate the lecturer's name
                if lecturer_name != lecturer_info.get("lecturer"):
                    st.error("Access denied: Lecturer's name does not match the database. Please try again.")
                    return

                # Final Step: Mark as Verified
                try:
                    for row in ws.iter_rows(min_row=2, values_only=False):
                        if row[3].value == selected_course and row[1].value == selected_date.strftime('%Y-%m-%d'):
                            row[4].value = f"Verified by {lecturer_name} and {st.session_state.rep_name}"  # Update the "Verified By" column
                    wb.save(ATTENDANCE_EXCEL_FILE)

                    

                    # Log the confirmation details
                    log_confirmation(
                        course=selected_course,
                        date=selected_date.strftime('%Y-%m-%d'),
                        time=datetime.now().strftime('%H:%M:%S'),
                        rep_name=st.session_state.rep_name,
                        role=st.session_state.selected_role,
                        lecturer_name=lecturer_name
                    )

                    st.success(f"Attendance records for {selected_course} on {selected_date.strftime('%Y-%m-%d')} successfully verified by {lecturer_name} and {st.session_state.rep_name}.")
                    st.session_state.verification_success = True
                except Exception as e:
                    st.error(f"An error occurred while marking attendance as verified: {e}")
                    st.session_state.verification_success = False

            # Display "Go Back" button after verification
            if st.session_state.verification_success:
                if st.button("Go Back"):
                    st.session_state.step = 1  # Reset to Step 1

    except FileNotFoundError:
        st.error("Attendance file not found.")
    except Exception as e:
        st.error(f"Error verifying attendance records: {e}")


def manual_attendance():
    st.subheader("Manual Attendance (For Online/Zoom Classes)")
    faces_db = load_data(FACE_DB_FILE, default={})
    if not faces_db:
        st.info("No registered students found.")
        return

    students = [details["name"] for details in faces_db.values()]
    courses = list(LECTURER_COURSE_MAPPING.keys())
    selected_course = st.selectbox("Select Course", courses)
    selected_date = st.date_input("Select Date", datetime.now().date())
    selected_students = st.multiselect("Select Students to Mark Present", students)

    if st.button("Mark Attendance"):
        try:
            wb = load_workbook(ATTENDANCE_EXCEL_FILE)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Date", "Time", "Course", "Verified By"])

        now = datetime.now()
        for student in selected_students:
            ws.append([student, selected_date.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'), selected_course, "Manual"])
        wb.save(ATTENDANCE_EXCEL_FILE)
        st.success(f"Attendance marked for: {', '.join(selected_students)} in {selected_course} on {selected_date.strftime('%Y-%m-%d')}.")

def remote_attendance():
    st.subheader("Remote/Online Attendance (Face Verification)")
    faces_db = load_data(FACE_DB_FILE, default={})
    if not faces_db:
        st.info("No registered students found.")
        return

    courses = list(LECTURER_COURSE_MAPPING.keys())
    selected_course = st.selectbox("Select Course", courses)
    st.info("Use your webcam to verify your face for online attendance.")

    captured_image = st.camera_input("Capture your face for attendance")
    verify_button = st.button("Verify and Mark Attendance")

    face_crop = None
    if captured_image:
        bytes_data = captured_image.getvalue()
        frame = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
        faces = detect_face(frame)
        if len(faces) == 1:
            x, y, w, h = faces[0]
            face_crop = frame[y:y + h, x:x + w]
            st.image(face_crop, channels="BGR", caption="Detected Face")
        else:
            st.warning("Please ensure only one face is visible in the frame.")

    if verify_button and face_crop is not None:
        try:
            embedding = DeepFace.represent(
                face_crop,
                model_name="ArcFace",
                detector_backend="retinaface",
                enforce_detection=True,
                align=True,
                anti_spoofing=False  # Set to True for production
            )[0]["embedding"]
        except Exception as e:
            st.error(f"Face verification failed: {e}")
            return

        # Matching logic
        best_match = None
        best_score = -1
        for person_id, data in faces_db.items():
            for emb in data.get("embeddings", []):
                score = cosine_similarity(embedding, emb)
                if score > best_score:
                    best_score = score
                    best_match = data["name"]
        if best_score > 0.5:
            success = mark_attendance(best_match, selected_course)
            if success:
                st.success(f"Attendance marked for {best_match} in {selected_course}.")
            else:
                st.error("Failed to mark attendance. Please try again.")
        else:
            st.warning("Face not recognized. Attendance not marked.")

# --- Admin Portal ---
def admin_login():
    """Admin authentication."""
    password = st.text_input("Enter Admin Password", type="password")
    if password:
        st.session_state['admin_attempted_login'] = True
        if password == ADMIN_PASSWORD:
            return True
        else:
            return False
    return False
    return False  # No password entered yet
def manage_confirmation_logs():
    """Allow admin to view and manage confirmation logs."""
    st.subheader("Manage Confirmation Logs")

    try:
        # Load the confirmation log file
        wb = load_workbook(CONFIRMATION_LOG_FILE)
        ws = wb.active

        # Extract confirmation logs
        logs = [
            {
                "Course": row[0].value,
                "Date": row[1].value,
                "Time": row[2].value,
                "Course Representative": row[3].value,
                "Role": row[4].value,
                "Lecturer": row[5].value,
                "Timestamp": row[6].value,
            }
            for row in ws.iter_rows(min_row=2, values_only=False)
        ]

        if not logs:
            st.info("No confirmation logs found.")
            return

        # Display logs in a table
        st.write("### Confirmation Logs")
        st.dataframe(logs)

        # Allow admin to remove a specific log
        st.write("### Remove a Confirmation Log")
        log_to_remove = st.selectbox(
            "Select Log to Remove",
            [f"{log['Course']} on {log['Date']} at {log['Time']}" for log in logs],
        )

        if st.button("Remove Log"):
            # Find and remove the selected log
            for row in ws.iter_rows(min_row=2, values_only=False):
                if (
                    row[0].value == log_to_remove.split(" on ")[0]
                    and row[1].value == log_to_remove.split(" on ")[1].split(" at ")[0]
                    and row[2].value == log_to_remove.split(" at ")[1]
                ):
                    ws.delete_rows(row[0].row)
                    wb.save(CONFIRMATION_LOG_FILE)
                    st.success(f"Log for {log_to_remove} removed successfully.")
                    return

    except FileNotFoundError:
        st.error("Confirmation log file not found.")
    except Exception as e:
        st.error(f"Error managing confirmation logs: {e}")

def manage_users():
    """Allow admin to manage lecturers, course representatives, and student credentials."""
    global LECTURER_COURSE_MAPPING  # Declare LECTURER_COURSE_MAPPING as global

    logged_in = admin_login()
    if not logged_in:
        if st.session_state.get('admin_attempted_login', False):
            st.warning("Access denied: Incorrect admin password.")
        return

    st.subheader("Admin Portal - Manage Users")
    users = load_data(USER_DB_FILE, {"Lecturers": {}, "CourseReps": {}})
    faces_db = load_data(FACE_DB_FILE, default={})  # Load registered faces

    menu = [
        "Add Lecturer",
        "Remove Lecturer",
        "Add Course Representative",
        "Remove Course Representative",
        "Update Students Credentials",
        "Update Lecturer Emails",
        "Manage Confirmation Logs", 
         "Regenerate All Face Embeddings"  # New option
    ]
    choice = st.selectbox("Select Action", menu)

    if choice == "Manage Confirmation Logs":
        manage_confirmation_logs()

    elif choice == "Add Lecturer":
        st.subheader("Add Lecturer")
        name = st.text_input("Lecturer's Name").strip()
        course = st.text_input("Enter Course Name (New or Existing)").strip()
        email = st.text_input("Lecturer's Email Address").strip()
        if st.button("Add Lecturer"):
            if not name or not email or not course:
                st.warning("All fields (name, email, and course) are required.")
                return

            # Update LECTURER_COURSE_MAPPING
            if course in LECTURER_COURSE_MAPPING:
                st.info(f"Course '{course}' already exists. Updating lecturer details.")
            else:
                st.info(f"Course '{course}' is new. Adding it to the system.")

            LECTURER_COURSE_MAPPING[course] = {"lecturer": name, "email": email}
            save_lecturers(LECTURER_COURSE_MAPPING)  # Save the updated data to the file
            st.success(f"Lecturer {name} assigned to course: {course}.")

    elif choice == "Remove Lecturer":
        st.subheader("Remove Lecturer")
        LECTURER_COURSE_MAPPING = load_lecturers()
        lecturer_options = [f"{details['lecturer']} ({course})" for course, details in LECTURER_COURSE_MAPPING.items()]
        if not lecturer_options:
            st.info("No lecturers found. Please add a lecturer first.")
            return

        selected_lecturer = st.selectbox("Select Lecturer to Remove", lecturer_options)
        if st.button("Remove Lecturer"):
            course_to_remove = selected_lecturer.split(" (")[-1][:-1]
            if course_to_remove in LECTURER_COURSE_MAPPING:
                del LECTURER_COURSE_MAPPING[course_to_remove]
                save_lecturers(LECTURER_COURSE_MAPPING)
                st.success(f"Lecturer for {course_to_remove} removed successfully.")
            else:
                st.warning(f"No lecturer found for {course_to_remove}.")

    elif choice == "Add Course Representative":
        st.subheader("Add Course Representative")
        name = st.text_input("Representative's Name").strip()
        course = st.selectbox("Course", list(LECTURER_COURSE_MAPPING.keys()))
        role = st.selectbox("Role", ["Head Course Rep", "Assistant Course Rep"])
        email = st.text_input("Representative's Email Address").strip()
        if st.button("Add"):
            if not name or not email:
                st.warning("Both name and email are required.")
                return
            if course not in users["CourseReps"]:
                users["CourseReps"][course] = []
            users["CourseReps"][course].append({"name": name, "email": email, "role": role})
            save_data(users, USER_DB_FILE)
            st.success(f"{role} {name} added for {course}.")

    elif choice == "Remove Course Representative":
        st.subheader("Remove Course Representative")
        if not users["CourseReps"]:
            st.info("No course representatives found. Please add one first.")
            return

        course_to_remove_from = st.selectbox("Select Course", list(users["CourseReps"].keys()))
        reps = users["CourseReps"][course_to_remove_from]
        rep_names = [f"{rep['name']} ({rep['role']})" for rep in reps]
        name_to_remove = st.selectbox("Select Representative to Remove", rep_names)

        if st.button("Remove"):
            name_part = name_to_remove.split(' ')[0]
            updated_reps = [rep for rep in reps if rep['name'] != name_part]
            users["CourseReps"][course_to_remove_from] = updated_reps
            save_data(users, USER_DB_FILE)
            st.success(f"Course Representative {name_part} removed from {course_to_remove_from}.")

    elif choice == "Update Students Credentials":
        st.subheader("Update Students Credentials")
        if not faces_db:
            st.info("No registered students found.")
            return
    
        student_id = st.selectbox("Select Student", list(faces_db.keys()))
        if student_id:
            student = faces_db[student_id]
            new_name = st.text_input("Update Student's Name", student["name"]).strip()
            new_email = st.text_input("Update Student's Email", student.get("email", "")).strip()
    
            st.info("Use the camera below to update the student's face images (optional, multiple angles recommended).")
            if "update_multi_face_images" not in st.session_state:
                st.session_state.update_multi_face_images = []
    
            captured_image = st.camera_input("Capture New Face Image (multiple angles recommended)", key="update_multi_frame")
    
            if captured_image:
                bytes_data = captured_image.getvalue()
                frame = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
                faces = detect_face(frame)
                if len(faces) == 1:
                    x, y, w, h = faces[0]
                    face_crop = frame[y:y + h, x:x + w]
                    st.session_state.update_multi_face_images.append(face_crop)
                    st.image(face_crop, channels="BGR", caption=f"Captured Face #{len(st.session_state.update_multi_face_images)}")
                else:
                    st.warning("Please ensure only one face is visible in the frame.")
    
            if st.button("Finish Face Update"):
                if len(st.session_state.update_multi_face_images) < 1:
                    st.warning("Please capture at least one face image.")
                else:
                    face_paths = []
                    face_embeddings = []
                    for idx, face_img in enumerate(st.session_state.update_multi_face_images):
                        face_path = os.path.join(FACE_DIR, f"{student_id}_update_{idx}.jpg")
                        cv2.imwrite(face_path, face_img)
                        face_paths.append(face_path)
                        try:
                            embedding = DeepFace.represent(
                                face_img,
                                model_name="ArcFace",
                                detector_backend="retinaface",
                                enforce_detection=True,
                                align=True,
                                anti_spoofing=False
                            )[0]["embedding"]
                            face_embeddings.append(embedding)
                        except ValueError as ve:
                            st.warning(f"Face image #{idx+1} was rejected: {ve}")
                        except Exception as e:
                            st.error(f"An unexpected error occurred for image #{idx+1}: {e}")
                    student["paths"] = face_paths
                    student["embeddings"] = face_embeddings
                    faces_db[student_id] = student
                    save_data(faces_db, FACE_DB_FILE)
                    st.success(f"Student {new_name}'s details and face images updated successfully!")
                    st.session_state.update_multi_face_images = []
    
            if st.button("Update Student"):
                if not new_name or not new_email:
                    st.warning("Both name and email are required.")
                    return
    
                student["name"] = new_name
                student["email"] = new_email
    
                if captured_image and frame is not None:
                    try:
                        faces = detect_face(frame)
                        if len(faces) == 0:
                            st.warning("No face detected. Please ensure the face is clearly visible.")
                            return
                        if len(faces) > 1:
                            st.warning("Multiple faces detected. Please ensure only one face is visible.")
                            return
    
                        x, y, w, h = faces[0]
                        face_crop = frame[y:y + h, x:x + w]
                        face_path = os.path.join(FACE_DIR, f"{student_id}.jpg")
                        cv2.imwrite(face_path, face_crop)
                        student["path"] = face_path
                        st.success("Face image updated successfully!")
                        st.image(face_crop, channels="BGR", caption="Updated Face")
                    except Exception as e:
                        st.error(f"Error updating face image: {e}")
    
                faces_db[student_id] = student
                save_data(faces_db, FACE_DB_FILE)
                st.success(f"Student {new_name}'s details updated successfully!")
    
            if st.button("Remove Student"):
                del faces_db[student_id]
                save_data(faces_db, FACE_DB_FILE)
                st.success(f"Student {student['name']} removed successfully!")


    elif choice == "Update Lecturer Emails":
        st.subheader("Update Lecturer Emails")
        course_to_update = st.selectbox("Select Course", list(LECTURER_COURSE_MAPPING.keys()))
        if course_to_update:
            lecturer_info = LECTURER_COURSE_MAPPING[course_to_update]
            current_email = lecturer_info.get("email", "Not Available")
            st.write(f"Current Email: {current_email}")
            new_email = st.text_input("Enter New Email Address").strip()
            if st.button("Update Email"):
                if not new_email:
                    st.warning("Email address cannot be empty.")
                    return
                LECTURER_COURSE_MAPPING[course_to_update]["email"] = new_email
                save_lecturers(LECTURER_COURSE_MAPPING)
                st.success(f"Email for {lecturer_info['lecturer']} updated successfully!")  

    def regenerate_all_embeddings():
        """Regenerate embeddings for all registered faces using robust settings."""
        faces_db = load_data(FACE_DB_FILE, default={})
        for person_id, details in faces_db.items():
            new_embeddings = []
            image_paths = details.get("paths", []) if "paths" in details else [details.get("path")]
            for img_path in image_paths:
                if img_path and os.path.exists(img_path):
                    try:
                        img = cv2.imread(img_path)
                        embedding = DeepFace.represent(
                            img,
                            model_name="ArcFace",
                            detector_backend="retinaface",
                            enforce_detection=True,
                            align=True,
                            anti_spoofing=False
                        )[0]["embedding"]
                        new_embeddings.append(embedding)
                    except Exception as e:
                        logging.error(f"Failed to regenerate embedding for {img_path}: {e}")
            faces_db[person_id]["embeddings"] = new_embeddings
        save_data(faces_db, FACE_DB_FILE)
        st.success("All face embeddings have been regenerated with robust settings.")

           

# --- Main Menu ---
def main_menu():
    st.title("Face Recognition Attendance System")
    menu = ["Register New Face", "Take Attendance", "View Attendance Records",
            "List Registered Faces", "Verify Attendance", "Daily Summary",
            "Manual Attendance","Remote Attendance", "Admin Portal", "Exit"]
    choice = st.sidebar.selectbox("Menu", menu)

    if choice == "Register New Face":
        register_face()
    elif choice == "Take Attendance":
        courses = list(LECTURER_COURSE_MAPPING.keys())
        course_to_take = st.selectbox("Select Course for Attendance", courses)
        verify_face(course_to_take)
    elif choice == "View Attendance Records":
        view_attendance()
    elif choice == "List Registered Faces":
        list_faces()
    elif choice == "Verify Attendance":
        verify_attendance()
    elif choice == "Daily Summary":
        generate_daily_summary()
    elif choice == "Manual Attendance":
        manual_attendance()
    elif choice == "Remote Attendance":
        remote_attendance()    
    elif choice == "Admin Portal":
        manage_users()
    elif choice == "Exit":
        st.info("Goodbye!")
    st.stop()
if __name__ == "__main__":
    main_menu()