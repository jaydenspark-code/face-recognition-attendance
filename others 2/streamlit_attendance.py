import streamlit as st
import cv2
import os
import pickle
import logging
from deepface import DeepFace
from datetime import datetime
from openpyxl import Workbook, load_workbook
import numpy as np  # Import numpy for image processing

# ... (Your existing logging and file configurations) ...
# ... (Your existing functions: save_faces, load_faces, detect_face, mark_attendance) ...

def register_face_streamlit():
    st.subheader("Register New Face")
    name = st.text_input("Enter the person's name:")
    captured_image = st.camera_input("Capture face")
    if captured_image and name:
        try:
            bytes_data = captured_image.getvalue()
            frame = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
            faces = detect_face(frame)
            if faces:
                x, y, w, h = faces[0]
                face_crop = frame[y:y+h, x:x+w]
                person_id = name.lower().replace(" ", "_")
                face_path = os.path.join(FACE_DIR, f"{person_id}.jpg")
                cv2.imwrite(face_path, face_crop)
                faces_db = load_faces()
                faces_db[person_id] = {"name": name, "path": face_path}
                save_faces(faces_db)
                st.success(f"Face for {name} registered successfully!")
                st.image(face_crop, caption="Registered Face", use_column_width=True)
            else:
                st.warning("No face detected. Please try again.")
                st.image(frame, caption="Webcam Feed", use_column_width=True)
        except Exception as e:
            st.error(f"An error occurred during registration: {e}")

def verify_face_streamlit():
    st.subheader("Verify Face (Mark Attendance)")
    faces = load_faces()
    if not faces:
        st.info("No faces registered yet.")
        return

    captured_image = st.camera_input("Show your face for verification")
    if captured_image:
        try:
            bytes_data = captured_image.getvalue()
            frame = cv2.imdecode(np.frombuffer(bytes_data, np.uint8), cv2.IMREAD_COLOR)
            face_rects = detect_face(frame)
            for (x, y, w, h) in face_rects:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                face_crop = frame[y:y + h, x:x + w]
                temp_path = "temp_face.jpg"
                cv2.imwrite(temp_path, face_crop)

                best_match = None
                best_confidence = float('inf')
                CONFIDENCE_THRESHOLD = 0.55

                for person_id, data in faces.items():
                    try:
                        result = DeepFace.verify(
                            img1_path=temp_path,
                            img2_path=data["path"],
                            detector_backend='opencv',
                            enforce_detection=False
                        )
                        if result.get("verified", False):
                            confidence = result.get("distance", float('inf'))
                            if confidence < CONFIDENCE_THRESHOLD and confidence < best_confidence:
                                best_confidence = confidence
                                best_match = data["name"]
                    except Exception as e:
                        st.error(f"Error verifying with {data['name']}: {e}")

                if best_match:
                    st.success(f"Attendance marked for {best_match}")
                    mark_attendance(best_match)
                else:
                    st.info("No match found.")
                os.remove(temp_path)
            st.image(frame, caption="Webcam Feed with Face Detection", use_column_width=True)
        except Exception as e:
            st.error(f"An error occurred during verification: {e}")

def list_faces_streamlit():
    st.subheader("Registered Faces")
    faces = load_faces()
    if faces:
        face_list = []
        for person_id, details in faces.items():
            face_list.append({"ID": person_id, "Name": details['name'], "Path": details['path']})
        st.dataframe(face_list)
    else:
        st.info("No registered faces found.")

def view_attendance_streamlit():
    st.subheader("Attendance Records")
    try:
        wb = load_workbook(ATTENDANCE_EXCEL_FILE)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        if data:
            st.dataframe(data, column_names=header)
        else:
            st.info("No attendance records found.")
    except FileNotFoundError:
        st.error("Attendance file not found.")
    except Exception as e:
        st.error(f"Error loading attendance: {e}")

def main():
    st.title("Face Recognition Attendance System")
    menu = ["Register New Face", "Verify Face (Mark Attendance)", "List Registered Faces", "View Attendance Records"]
    choice = st.sidebar.selectbox("Menu", menu)

    if choice == "Register New Face":
        register_face_streamlit()
    elif choice == "Verify Face (Mark Attendance)":
        verify_face_streamlit()
    elif choice == "List Registered Faces":
        list_faces_streamlit()
    elif choice == "View Attendance Records":
        view_attendance_streamlit()

if __name__ == "__main__":
    main()