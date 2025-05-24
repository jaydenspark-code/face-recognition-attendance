import pandas as pd
import openpyxl
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Sample data
patients_data = [
    {"PatientID": 1, "Name": "John Doe", "DOB": "1980-01-01", "Gender": "M", "AdmissionDate": "2024-06-01", "DischargeDate": "2024-06-05"},
    {"PatientID": 2, "Name": "Jane Smith", "DOB": "1990-02-15", "Gender": "F", "AdmissionDate": "2024-06-03", "DischargeDate": "2024-06-07"},
    {"PatientID": 3, "Name": "Alice Brown", "DOB": "1975-07-20", "Gender": "F", "AdmissionDate": "2024-06-04", "DischargeDate": ""},
]

if os.path.exists("MIS_Patient_Management.xlsx"):
    df_doctors = pd.read_excel("MIS_Patient_Management.xlsx", sheet_name="Doctors")
else:
    doctors_data = [
        {"DoctorID": 1, "Name": "Dr. Adams", "Specialty": "Cardiology"},
        {"DoctorID": 2, "Name": "Dr. Baker", "Specialty": "Neurology"},
        {"DoctorID": 3, "Name": "Dr. Clark", "Specialty": "General"},
    ]
    df_doctors = pd.DataFrame(doctors_data)

appointments_data = [
    {"AppointmentID": 1, "PatientID": 1, "DoctorID": 1, "Date": "2024-06-02", "Status": "Completed"},
    {"AppointmentID": 2, "PatientID": 2, "DoctorID": 2, "Date": "2024-06-04", "Status": "Completed"},
    {"AppointmentID": 3, "PatientID": 3, "DoctorID": 3, "Date": "2024-06-05", "Status": "Scheduled"},
]

# Create DataFrames
df_patients = pd.DataFrame(patients_data)
df_appointments = pd.DataFrame(appointments_data)

# Create workbook and sheets
wb = Workbook()
ws_patients = wb.active
ws_patients.title = "Patients"
ws_doctors = wb.create_sheet("Doctors")
ws_appointments = wb.create_sheet("Appointments")
ws_dashboard = wb.create_sheet("Dashboard")

# Write data to sheets
for r in dataframe_to_rows(df_patients, index=False, header=True):
    ws_patients.append(r)
for r in dataframe_to_rows(df_doctors, index=False, header=True):
    ws_doctors.append(r)
for r in dataframe_to_rows(df_appointments, index=False, header=True):
    ws_appointments.append(r)

# Dashboard KPIs
ws_dashboard["A1"] = "Key Performance Indicators"
ws_dashboard["A3"] = "Total Patients"
ws_dashboard["B3"] = f"=COUNTA(Patients!A2:A1000)"
ws_dashboard["A4"] = "Currently Admitted"
ws_dashboard["B4"] = '=COUNTIFS(Patients!F2:F4,"")'
ws_dashboard["A5"] = "Total Appointments"
ws_dashboard["B5"] = f"=COUNTA(Appointments!A2:A1000)"
ws_dashboard["A6"] = "Completed Appointments"
ws_dashboard["B6"] = f'=COUNTIFS(Appointments!E2:E1000,"Completed")'

# Chart: Appointments per Doctor
ws_dashboard["A8"] = "Appointments per Doctor"
ws_dashboard["A9"] = "Doctor"
ws_dashboard["B9"] = "Appointments"
for idx, doc in enumerate(df_doctors["Name"], start=10):
    ws_dashboard[f"A{idx}"] = doc
    ws_dashboard[f"B{idx}"] = f'=COUNTIFS(Appointments!C2:C1000,Doctors!A{idx-8})'

chart = BarChart()
chart.title = "Appointments per Doctor"
chart.x_axis.title = "Doctor"
chart.y_axis.title = "Appointments"
data = Reference(ws_dashboard, min_col=2, min_row=9, max_row=9+len(df_doctors))
cats = Reference(ws_dashboard, min_col=1, min_row=10, max_row=9+len(df_doctors))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_dashboard.add_chart(chart, "D9")

# Save file
wb.save("MIS_Patient_Management.xlsx")